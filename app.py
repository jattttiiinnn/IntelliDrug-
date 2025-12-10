import streamlit as st
# Trigger reload
import asyncio
import json
import hashlib
from datetime import datetime
from master_agent import MasterAgent
import pandas as pd
import plotly.graph_objects as go
from pathlib import Path
import base64
import os
import re
import tempfile
from typing import List, Tuple, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from visualization_components import (
    create_risk_radar,
    create_timeline_gantt,
    create_market_funnel,
    export_dashboard
)

# ======================================================================================
# HELPER FUNCTIONS FOR EXCEL EXPORT
# ======================================================================================

def create_excel_export(all_results: Dict[str, Any], molecule_name: str, disease_name: str) -> str:
    """
    Create an Excel workbook with analysis results.
    
    Args:
        all_results: Dictionary containing all analysis results
        molecule_name: Name of the molecule
        disease_name: Name of the disease
        
    Returns:
        Path to the generated Excel file
    """
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Executive Summary
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    
    # Add headers
    headers = ["Molecule", "Disease", "Date", "Recommendation", "Confidence"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws_summary.column_dimensions[get_column_letter(col_num)].width = 25
    
    # Add data
    synthesis = all_results.get('synthesis', {})
    ws_summary.append([
        molecule_name,
        disease_name,
        synthesis.get('timestamp', '').split('T')[0],
        synthesis.get('recommendation', 'N/A'),
        synthesis.get('confidence_display', 'N/A')
    ])
    
    # Format recommendation cell
    rec_cell = ws_summary.cell(row=2, column=4)
    rec_value = str(rec_cell.value or '').lower()
    if 'proceed' in rec_value and 'caution' not in rec_value:
        rec_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    elif 'caution' in rec_value:
        rec_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    elif 'not recommend' in rec_value or 'reject' in rec_value:
        rec_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Add a space
    ws_summary.append([])
    
    # Add key factors
    ws_summary.append(["Key Factors:"])
    for factor in synthesis.get('key_factors', [])[:5]:  # Limit to top 5 key factors
        ws_summary.append(["", factor])
    
    # Add strengths and weaknesses
    ws_summary.append([])
    ws_summary.append(["Strengths:"])
    for strength in synthesis.get('strengths', [])[:5]:  # Limit to top 5
        ws_summary.append(["", strength])
    
    ws_summary.append([])
    ws_summary.append(["Weaknesses:"])
    for weakness in synthesis.get('weaknesses', [])[:5]:  # Limit to top 5
        ws_summary.append(["", weakness])
    
    # Create sheets for each agent
    agent_order = [
        'patent_analysis', 'clinical_analysis', 'market_analysis',
        'web_analysis', 'exim_analysis', 'internal_analysis'
    ]
    
    for agent_name in agent_order:
        agent_data = all_results.get(agent_name, {})
        if not agent_data or not isinstance(agent_data, dict):
            continue
            
        # Create sheet for agent
        ws = wb.create_sheet(title=agent_name.replace('_', ' ').title())
        
        # Add headers
        headers = ["Finding", "Evidence Strength", "Sources", "Confidence"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_num)].width = 30
        
        # Add findings
        findings = agent_data.get('findings', [])
        if not isinstance(findings, list):
            findings = []
            
        for i, finding in enumerate(findings, 2):
            if not isinstance(finding, dict):
                continue
                
            # Get finding text
            finding_text = finding.get('finding') or finding.get('description', '')
            if not isinstance(finding_text, str):
                finding_text = str(finding_text)
            
            # Get evidence strength (normalize to 0-100)
            strength = finding.get('evidence_strength', 50)
            try:
                strength = min(100, max(0, int(float(strength))))
            except (ValueError, TypeError):
                strength = 50
            
            # Get sources
            sources = finding.get('sources', [])
            if not isinstance(sources, list):
                sources = []
            sources_text = "; ".join(str(s) for s in sources[:3])  # Limit to 3 sources
            
            # Get confidence
            confidence = finding.get('confidence', 0)
            try:
                confidence = min(100, max(0, int(float(confidence))))
            except (ValueError, TypeError):
                confidence = 0
            
            # Add row
            ws.append([finding_text, strength, sources_text, f"{confidence}%"])
            
            # Apply conditional formatting to evidence strength
            strength_cell = ws.cell(row=i, column=2)
            if strength >= 70:
                strength_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif strength >= 30:
                strength_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                strength_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # Apply conditional formatting to confidence
            conf_cell = ws.cell(row=i, column=4)
            if confidence >= 70:
                conf_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif confidence >= 30:
                conf_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                conf_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 50)  # Cap at 50
    
    # Create Risk Matrix sheet
    ws_risk = wb.create_sheet(title="Risk Matrix")
    
    # Add headers
    headers = ["Risk Factor", "Severity", "Likelihood", "Risk Score", "Mitigation"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_risk.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws_risk.column_dimensions[get_column_letter(col_num)].width = 25
    
    # Add risks from synthesis
    risks = synthesis.get('risks', [])
    if not risks or (isinstance(risks, list) and len(risks) == 1 and "No significant risks" in str(risks[0])):
        ws_risk.append(["No significant risks identified", "", "", "", ""])
    else:
        for risk in risks:
            if not isinstance(risk, dict):
                risk = {"description": str(risk), "severity": "Medium", "likelihood": "Medium", "mitigation": ""}
            
            severity = risk.get("severity", "Medium")
            likelihood = risk.get("likelihood", "Medium")
            
            # Calculate risk score (simple multiplication)
            severity_scores = {"Low": 1, "Medium": 2, "High": 3}
            likelihood_scores = {"Low": 1, "Medium": 2, "High": 3}
            
            score = severity_scores.get(str(severity).capitalize(), 1) * \
                   likelihood_scores.get(str(likelihood).capitalize(), 1)
            
            ws_risk.append([
                risk.get("description", ""),
                severity,
                likelihood,
                score,
                risk.get("mitigation", "")
            ])
            
            # Color code risk score
            risk_cell = ws_risk.cell(row=ws_risk.max_row, column=4)
            if score >= 6:
                risk_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif score >= 3:
                risk_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                risk_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    # Auto-adjust column widths for risk matrix
    for col in ws_risk.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_risk.column_dimensions[column].width = min(adjusted_width, 40)  # Cap at 40
    
    # Save to a temporary file
    temp_dir = tempfile.gettempdir()
    filename = f"{molecule_name.replace(' ', '_')}_{disease_name.replace(' ', '_')}_analysis.xlsx"
    filepath = os.path.join(temp_dir, filename)
    wb.save(filepath)
    
    return filepath

# ======================================================================================
# HELPER FUNCTIONS FOR SAVE/LOAD
# ======================================================================================
SAVED_ANALYSES_DIR = Path("saved_analyses")
SAVED_ANALYSES_DIR.mkdir(exist_ok=True)

def generate_analysis_id(molecule_name: str, disease_name: str) -> str:
    """Generate a unique ID for an analysis."""
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    unique_str = f"{molecule_name}_{disease_name}_{timestamp}"
    return hashlib.md5(unique_str.encode()).hexdigest()

def save_analysis(analysis_data: dict) -> str:
    """Save analysis data to a JSON file."""
    analysis_id = analysis_data["id"]
    filepath = SAVED_ANALYSES_DIR / f"{analysis_id}.json"
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(analysis_data, f, ensure_ascii=False, indent=2)
    return analysis_id

def load_analysis(analysis_id: str) -> dict:
    """Load analysis data from a JSON file."""
    filepath = SAVED_ANALYSES_DIR / f"{analysis_id}.json"
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None

def list_saved_analyses() -> list:
    """List all saved analyses with metadata."""
    analyses = []
    for filepath in SAVED_ANALYSES_DIR.glob("*.json"):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                analyses.append({
                    'id': data['id'],
                    'molecule': data['molecule_name'],
                    'disease': data['disease_name'],
                    'timestamp': datetime.fromtimestamp(int(data['timestamp'])).strftime('%Y-%m-%d %H:%M:%S'),
                    'recommendation': data.get('results', {}).get('recommendation', 'No recommendation')
                })
        except (json.JSONDecodeError, KeyError):
            continue
    
    # Sort by timestamp (newest first)
    return sorted(analyses, key=lambda x: x['timestamp'], reverse=True)


def find_similar_molecules(molecule_name: str, disease_name: str) -> List[Tuple[str, str]]:
    """
    Find molecules similar to the given molecule using web search.
    
    Args:
        molecule_name: Name of the molecule to find similar ones for
        disease_name: Name of the disease for context
        
    Returns:
        List of tuples (molecule_name, similarity_reason)
    """
    try:
        # Initialize web intelligence agent if not already done
        if 'web_agent' not in st.session_state:
            from web_intelligence_agent import WebIntelligenceAgent
            st.session_state.web_agent = WebIntelligenceAgent()
        
        # Search for similar molecules
        query1 = f"molecules similar to {molecule_name} same class mechanism target"
        query2 = f"alternatives to {molecule_name} for {disease_name}"
        
        results1 = st.session_state.web_agent.search(query1, num_results=3)
        results2 = st.session_state.web_agent.search(query2, num_results=3)
        
        # Extract and deduplicate molecules
        molecules = []
        
        # Process first query results (similar molecules)
        for result in results1:
            # Simple regex to find potential molecule names (uppercase words with numbers/letters)
            mols = re.findall(r'\b[A-Z][a-zA-Z0-9]+\b', result['snippet'])
            molecules.extend([(m, "same_class") for m in mols if len(m) > 3])  # Filter out short words
        
        # Process second query results (alternatives)
        for result in results2:
            mols = re.findall(r'\b[A-Z][a-zA-Z0-9]+\b', result['snippet'])
            molecules.extend([(m, "alternative") for m in mols if len(m) > 3])
        
        # Deduplicate while preserving order
        seen = set()
        unique_molecules = []
        for mol, mol_type in molecules:
            if mol != molecule_name and mol not in seen:
                seen.add(mol)
                unique_molecules.append((mol, mol_type))
        
        return unique_molecules[:5]  # Return max 5 molecules
        
    except Exception as e:
        st.error(f"Error finding similar molecules: {str(e)}")
        return []


# ======================================================================================
# PAGE CONFIGURATION
# ======================================================================================
st.set_page_config(
    page_title="IntelliDrug AI - Drug Repurposing",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ======================================================================================
# CUSTOM CSS - MODERN DESIGN WITH ANIMATIONS
# ======================================================================================
st.markdown("""
<style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');
    
    /* ==================== GLOBAL STYLES ==================== */
    * {
        font-family: 'Inter', sans-serif;
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stDeployButton {display: none;}
    
    /* Remove default padding */
    .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
        max-width: 1400px;
    }
    
    /* ==================== ANIMATED GRADIENT BACKGROUND ==================== */
    .stApp {
        background: linear-gradient(135deg, #0f0c29, #302b63, #24243e);
        background-size: 400% 400%;
        animation: gradientShift 15s ease infinite;
    }
    
    @keyframes gradientShift {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }
    
    /* ==================== FLOATING PARTICLES BACKGROUND ==================== */
    .stApp::before {
        content: '';
        position: fixed;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        background-image: 
            radial-gradient(2px 2px at 20% 30%, rgba(78, 205, 196, 0.4), transparent),
            radial-gradient(2px 2px at 60% 70%, rgba(78, 205, 196, 0.3), transparent),
            radial-gradient(3px 3px at 50% 50%, rgba(78, 205, 196, 0.4), transparent),
            radial-gradient(2px 2px at 80% 10%, rgba(78, 205, 196, 0.3), transparent),
            radial-gradient(2px 2px at 90% 60%, rgba(78, 205, 196, 0.4), transparent),
            radial-gradient(3px 3px at 30% 80%, rgba(78, 205, 196, 0.3), transparent),
            radial-gradient(2px 2px at 15% 90%, rgba(78, 205, 196, 0.4), transparent),
            radial-gradient(3px 3px at 75% 25%, rgba(78, 205, 196, 0.3), transparent),
            radial-gradient(2px 2px at 40% 60%, rgba(78, 205, 196, 0.4), transparent),
            radial-gradient(2px 2px at 95% 85%, rgba(78, 205, 196, 0.3), transparent);
        background-size: 200% 200%;
        animation: particleFloat 25s ease-in-out infinite;
        pointer-events: none;
        z-index: 0;
        opacity: 0.5;
    }

    @keyframes particleFloat {
        0%, 100% { 
            background-position: 0% 0%, 100% 100%, 50% 50%, 100% 0%, 0% 100%, 75% 25%, 25% 75%, 60% 40%, 40% 60%, 90% 10%;
            transform: translateY(0);
        }
        50% { 
            background-position: 100% 100%, 0% 0%, 75% 75%, 50% 50%, 100% 0%, 25% 75%, 75% 25%, 40% 60%, 60% 40%, 10% 90%;
            transform: translateY(-20px);
        }
    }
    
    /* ==================== HERO SECTION ==================== */
    .hero-section {
        text-align: center;
        padding: 60px 20px;
        margin-bottom: 40px;
        position: relative;
        z-index: 1;
    }
    
    .hero-title {
        font-size: 4rem;
        font-weight: 700;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 10px;
        animation: fadeInDown 1s ease;
    }
    
    .hero-subtitle {
        font-size: 1.5rem;
        color: rgba(255, 255, 255, 0.8);
        font-weight: 300;
        animation: fadeInUp 1s ease 0.2s both;
    }
    
    .hero-description {
        font-size: 1.1rem;
        color: rgba(255, 255, 255, 0.6);
        margin-top: 15px;
        animation: fadeInUp 1s ease 0.4s both;
    }
    
    @keyframes fadeInDown {
        from {
            opacity: 0;
            transform: translateY(-30px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(30px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    /* ==================== GLASSMORPHISM CARDS ==================== */
    .glass-card {
        background: rgba(255, 255, 255, 0.05);
        backdrop-filter: blur(10px);
        border-radius: 20px;
        border: 1px solid rgba(255, 255, 255, 0.1);
        padding: 30px;
        margin-bottom: 20px;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
        transition: all 0.3s ease;
        animation: fadeIn 0.8s ease;
    }
    
    .glass-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 12px 40px 0 rgba(78, 205, 196, 0.2);
        border-color: rgba(78, 205, 196, 0.3);
    }
    
    @keyframes fadeIn {
        from { opacity: 0; transform: scale(0.95); }
        to { opacity: 1; transform: scale(1); }
    }
    
    /* ==================== INPUT FIELDS ==================== */
    .stTextInput > div > div > input {
        background: rgba(255, 255, 255, 0.08);
        border: 2px solid rgba(255, 255, 255, 0.1);
        border-radius: 12px;
        color: white;
        font-size: 16px;
        padding: 15px 20px;
        transition: all 0.3s ease;
    }
    
    .stTextInput > div > div > input:focus {
        background: rgba(255, 255, 255, 0.12);
        border-color: #4ECDC4;
        box-shadow: 0 0 20px rgba(78, 205, 196, 0.4);
        transform: scale(1.02);
    }
    
    .stTextInput > label {
        color: rgba(255, 255, 255, 0.9) !important;
        font-weight: 600;
        font-size: 1.1rem;
        margin-bottom: 8px;
    }
    
    /* ==================== BUTTONS ==================== */
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 15px 40px;
        font-size: 18px;
        font-weight: 600;
        width: 100%;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
        cursor: pointer;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 25px rgba(102, 126, 234, 0.6);
        background: linear-gradient(135deg, #764ba2 0%, #667eea 100%);
    }
    
    .stButton > button:active {
        transform: translateY(0);
    }
    
    /* ==================== PROGRESS SECTION ==================== */
    .agent-status {
        background: rgba(255, 255, 255, 0.05);
        backdrop-filter: blur(10px);
        border-radius: 15px;
        padding: 20px;
        margin: 10px 0;
        border-left: 4px solid #4ECDC4;
        transition: all 0.3s ease;
        animation: slideInLeft 0.5s ease;
    }
    
    @keyframes slideInLeft {
        from {
            opacity: 0;
            transform: translateX(-30px);
        }
        to {
            opacity: 1;
            transform: translateX(0);
        }
    }
    
    .agent-status.running {
        border-left-color: #ffc107;
        animation: pulse 2s ease infinite;
    }
    
    .agent-status.complete {
        border-left-color: #28a745;
    }
    
    .agent-status.failed {
        border-left-color: #dc3545;
    }
    
    @keyframes pulse {
        0%, 100% { opacity: 1; }
        50% { opacity: 0.7; }
    }
    
    /* ==================== PROGRESS BAR ==================== */
    .stProgress > div > div > div {
        background: linear-gradient(90deg, #4ECDC4, #44A08D);
        border-radius: 10px;
    }
    
    /* ==================== TABS ==================== */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: rgba(255, 255, 255, 0.05);
        padding: 10px;
        border-radius: 12px;
    }
    
    .stTabs [data-baseweb="tab"] {
        background: transparent;
        color: rgba(255, 255, 255, 0.7);
        border-radius: 8px;
        padding: 12px 24px;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background: rgba(255, 255, 255, 0.1);
        color: white;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white !important;
    }
    
    /* ==================== EVIDENCE BADGES ==================== */
    .evidence-badge {
        display: inline-block;
        padding: 0.25em 0.6em;
        font-size: 0.8em;
        font-weight: 600;
        line-height: 1.5;
        text-align: center;
        white-space: nowrap;
        vertical-align: baseline;
        border-radius: 0.25rem;
        margin-left: 0.5rem;
    }
    .evidence-strong {
        background-color: #d4edda;
        color: #155724;
        border: 1px solid #c3e6cb;
    }
    .evidence-moderate {
        background-color: #fff3cd;
        color: #856404;
        border: 1px solid #ffeeba;
    }
    .evidence-weak {
        background-color: #f8d7da;
        color: #721c24;
        border: 1px solid #f5c6cb;
    }
    
    /* ==================== EXPANDER ==================== */
    .streamlit-expanderHeader {
        background: rgba(255, 255, 255, 0.05);
        border-radius: 10px;
        padding: 15px;
        color: white;
        font-weight: 600;
    }
    
    .streamlit-expanderHeader:hover {
        background: rgba(255, 255, 255, 0.1);
    }
    
    /* ==================== SUCCESS/WARNING/ERROR MESSAGES ==================== */
    .stSuccess, .stWarning, .stError, .stInfo {
        border-radius: 12px;
        backdrop-filter: blur(10px);
    }
</style>
""", unsafe_allow_html=True)

# ======================================================================================
# HERO SECTION
# ======================================================================================
st.markdown("""
<div class="hero-section">
    <div class="hero-title">🧬 IntelliDrug AI</div>
    <div class="hero-subtitle">Accelerating Drug Discovery with Multi-Agent Intelligence</div>
    <div class="hero-description">Transform 90 days of research into 4 hours with AI-powered analysis</div>
</div>
""", unsafe_allow_html=True)

# ======================================================================================
# SIDEBAR
# ======================================================================================
st.sidebar.header("Settings")

# Load saved analyses for the dropdown
saved_analyses = list_saved_analyses()

# Saved Analyses Section
with st.sidebar.expander("💾 Saved Analyses", expanded=True):
    if saved_analyses:
        # Dropdown to select a saved analysis
        analysis_options = ["Select an analysis..."] + [
            f"{a['molecule']} - {a['disease']} ({a['timestamp']})" 
            for a in saved_analyses
        ]
        
        selected_analysis = st.selectbox(
            "Load Analysis",
            options=analysis_options,
            key="selected_analysis"
        )
        
        # Handle loading of selected analysis
        if selected_analysis != "Select an analysis...":
            selected_idx = analysis_options.index(selected_analysis) - 1
            if 0 <= selected_idx < len(saved_analyses):
                analysis_id = saved_analyses[selected_idx]['id']
                loaded_analysis = load_analysis(analysis_id)
                
                if loaded_analysis:
                    st.session_state.loaded_analysis = loaded_analysis
                    st.session_state.molecule_name = loaded_analysis["molecule_name"]
                    st.session_state.disease_name = loaded_analysis["disease_name"]
                    st.success(f"✅ Loaded analysis for {loaded_analysis['molecule_name']} - {loaded_analysis['disease_name']}")
                    
                    if st.button("🔄 Re-run Analysis"):
                        st.session_state.pop("loaded_analysis", None)
                        st.rerun()
    else:
        st.info("No saved analyses found. Complete an analysis to save it here.")

st.sidebar.markdown("---")
st.sidebar.markdown("### Quick Stats")
st.sidebar.metric("Analyses this month", f"{len(saved_analyses)}")
st.sidebar.markdown("---")
st.sidebar.markdown("### About")
st.sidebar.info(
    "IntelliDrug AI leverages multiple intelligent agents "
    "to analyze patent, clinical, market, and internal company data."
)

# ======================================================================================
# ANALYSIS MODE SELECTION
# ======================================================================================
analysis_mode = st.radio(
    "Select Analysis Mode:",
    ["Single Molecule", "Compare Molecules"],
    horizontal=True
)

# ======================================================================================
# INPUT SECTION - GLASSMORPHISM CARDS
# ======================================================================================
if analysis_mode == "Single Molecule":
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### 💊 Molecule Name")
        molecule_name = st.text_input(
            "Molecule Name",
            value="Metformin",
            placeholder="e.g., Metformin, Aspirin, Semaglutide",
            label_visibility="collapsed"
        )
    with col2:
        st.markdown("### 🦠 Disease Target")
        disease_name = st.text_input(
            "Disease Name",
            value="NASH",
            placeholder="e.g., NASH, Diabetes, Cancer",
            label_visibility="collapsed"
        )
    analyze_button = st.button("▶ Run Comprehensive Analysis", use_container_width=True)
    molecules = [molecule_name] if molecule_name else []
    st.markdown('</div>', unsafe_allow_html=True)
else:  # Compare Molecules
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.markdown("### Enter 2-3 molecules to compare")
    col1, col2, col3 = st.columns(3)
    with col1:
        mol1 = st.text_input("Molecule 1", value="Metformin")
    with col2:
        mol2 = st.text_input("Molecule 2", value="Empagliflozin")
    with col3:
        mol3 = st.text_input("Molecule 3 (optional)", value="")
    disease_name = st.text_input("Target Disease", value="NASH")
    analyze_button = st.button("Compare Molecules")
    molecules = [m for m in [mol1, mol2, mol3] if m.strip()]
    if len(molecules) < 2:
        st.warning("Please enter at least 2 molecules to compare")
    st.markdown("</div>", unsafe_allow_html=True)

# ======================================================================================
# INITIALIZE MASTER AGENT
# ======================================================================================
master_agent = MasterAgent()

# Initialize session state for loaded analysis
if "loaded_analysis" not in st.session_state:
    st.session_state.loaded_analysis = None

# ======================================================================================
# RUN ANALYSIS
# ======================================================================================
if analyze_button or st.session_state.get("loaded_analysis"):
    # Use loaded analysis data if available
    if st.session_state.loaded_analysis:
        loaded_data = st.session_state.loaded_analysis
        molecule_name = loaded_data["molecule_name"]
        disease_name = loaded_data["disease_name"]
        all_results = loaded_data["results"]
        analysis_mode = "Single Molecule"  # Default to single molecule for loaded analyses
        
        st.markdown(f"""
        <div class="glass-card" style="text-align: center;">
            <h3 style="margin: 0; color: white;">🔍 Viewing Saved Analysis: {molecule_name} for {disease_name}</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # Add save button for the loaded analysis
        if st.button("💾 Save as New Analysis"):
            new_analysis_id = generate_analysis_id(molecule_name, disease_name)
            analysis_data = {
                "id": new_analysis_id,
                "molecule_name": molecule_name,
                "disease_name": disease_name,
                "timestamp": str(int(datetime.now().timestamp())),
                "results": all_results
            }
            save_analysis(analysis_data)
            st.success(f"✅ Analysis saved with ID: {new_analysis_id}")
            
    # Run new analysis if no loaded analysis or re-run was requested
    elif (analysis_mode == "Single Molecule" and (not molecule_name or not disease_name)) or \
         (analysis_mode == "Compare Molecules" and (len(molecules) < 2 or not disease_name)):
        st.warning("⚠️ Please fill in all required fields.")
    else:
        # ======================================================================================
        # PROGRESS SECTION
        # ======================================================================================
        st.markdown(f"""
        <div class="glass-card" style="text-align: center;">
            <h3 style="margin: 0; color: white;">🔬 Analyzing {molecule_name if analysis_mode == "Single Molecule" else "Multiple Molecules"} for {disease_name}</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # Create placeholders for agent status
        progress_placeholders = {}
        for agent_key in master_agent.agents.keys():
            progress_placeholders[agent_key] = st.empty()
        
                # Progress bar
        progress_bar = st.progress(0)
        
        if 'show_research_mode' not in st.session_state:
            st.session_state.show_research_mode = False

        # Add a button to toggle Research Mode
        if st.button("🔬 Show Research Mode", key="toggle_research_mode"):
            st.session_state.show_research_mode = not st.session_state.show_research_mode

        # If Research Mode is enabled, show the comparison
        if st.session_state.show_research_mode and analysis_mode == "Single Molecule":
            st.markdown("### 🧪 Research Mode: Analysis Comparison")
            st.markdown("Compare how different analytical approaches affect the recommendation:")
            
            # Create three columns for the comparison
            col1, col2, col3 = st.columns(3)
            
            with st.spinner("Running alternative analyses..."):
                from master_agent import AnalysisStrategy
                
                # Run all three strategies in parallel
                strategies = [
                    ("Standard", AnalysisStrategy.STANDARD),
                    ("Optimistic", AnalysisStrategy.OPTIMISTIC),
                    ("Conservative", AnalysisStrategy.CONSERVATIVE)
                ]
                
                # Run analyses
                # Run analyses
                async def run_comparison_strategies():
                    strategy_results = {}
                    for name, strategy in strategies:
                        strategy_results[name] = await master_agent.analyze_with_strategy(
                            molecule_name, disease_name, strategy
                        )
                    return strategy_results

                results = asyncio.run(run_comparison_strategies())
                
                # Display results in columns
                for i, (name, result) in enumerate(results.items()):
                    with [col1, col2, col3][i]:
                        synthesis = result.get("synthesis", {})
                        st.markdown(f"#### {name} Approach")
                        
                        # Style based on recommendation
                        rec = synthesis.get("recommendation", "").upper()
                        rec_style = "color: #4CAF50;"  # green
                        if "CAUTION" in rec:
                            rec_style = "color: #FFC107;"  # yellow
                        elif "NOT RECOMMENDED" in rec or "REJECT" in rec:
                            rec_style = "color: #F44336;"  # red
                        
                        st.markdown(
                            f'<div style="font-size: 1.2em; {rec_style} font-weight: bold; margin: 10px 0;">'
                            f'{synthesis.get("recommendation", "No recommendation")}'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                        
                        # Show confidence
                        st.markdown(f"**Confidence:** {synthesis.get('confidence_display', 'N/A')}")
                        
                        # Show key differences
                        key_factors = synthesis.get("key_factors", [])
                        if key_factors:
                            st.markdown("**Key Factors:**")
                            for factor in key_factors[:3]:  # Show top 3
                                st.markdown(f"- {factor}")
                        
                        # Show risks if conservative
                        if name == "Conservative" and "risks" in synthesis:
                            risks = synthesis["risks"]
                            if isinstance(risks, list) and len(risks) > 0:
                                st.markdown("**Risks Considered:**")
                                for risk in risks[:2]:  # Show top 2 risks
                                    if isinstance(risk, str):
                                        st.markdown(f"- ⚠️ {risk}")
                                    elif isinstance(risk, dict) and "description" in risk:
                                        st.markdown(f"- ⚠️ {risk['description']}")

            # Add explanation
            st.markdown("""
            <div style="margin-top: 20px; padding: 15px; background: rgba(30, 30, 30, 0.2); border-radius: 10px;">
                <h4>🔍 Understanding the Different Approaches</h4>
                <p>This comparison shows how different analytical perspectives can lead to different conclusions:</p>
                <ul>
                    <li><b>Standard:</b> Balanced approach with equal consideration of all factors</li>
                    <li><b>Optimistic:</b> Emphasizes potential benefits and market opportunities</li>
                    <li><b>Conservative:</b> Gives more weight to potential risks and challenges</li>
                </ul>
                <p>This demonstrates how assumptions and weightings can affect AI-driven recommendations.</p>
            </div>
            """, unsafe_allow_html=True)

        # ======================================================================================
        # ASYNC ANALYSIS WITH REAL-TIME UPDATES
        # ======================================================================================
        async def run_analysis():
            """Run analysis asynchronously and update progress."""
            self = master_agent  # Access master_agent in the local scope
            
            # Initialize progress
            for name in self.agents:
                self.progress[name] = "Pending"
            
            # Start analysis based on mode
            if analysis_mode == "Single Molecule":
                analysis_task = asyncio.create_task(
                    master_agent.analyze_repurposing_async(molecule_name, disease_name)
                )
            else:  # Compare Molecules
                analysis_task = asyncio.create_task(
                    master_agent.compare_molecules_async(molecules, disease_name)
                )
            
            # Update progress until analysis completes
            while not analysis_task.done():
                # Update progress bar
                completed = sum(1 for s in self.progress.values() if s == "Complete")
                total = len(self.progress)
                progress_bar.progress(min(100, int((completed / total) * 100)))
                
                # Update status text and agent boxes
                status_text = st.empty()
                status_text.text(f"Analyzing... ({completed}/{total} agents complete)")
                for name, ph in progress_placeholders.items():
                    state = self.progress.get(name, "Pending")
                    if state == "Complete":
                        emoji = "✅"
                    elif state == "Running":
                        emoji = "⏳"
                    elif state == "Failed":
                        emoji = "❌"
                    else:
                        emoji = "⏸️"
                    ph.markdown(f"**{name.replace('_',' ').title()}** {emoji} {state}")
                
                await asyncio.sleep(0.5)
            
            # Get final results
            return await analysis_task

        # Run analysis
        all_results = asyncio.run(run_analysis())
        
        # Save the analysis automatically
        if "loaded_analysis" not in st.session_state:  # Only auto-save new analyses
            analysis_id = generate_analysis_id(molecule_name, disease_name)
            analysis_data = {
                "id": analysis_id,
                "molecule_name": molecule_name,
                "disease_name": disease_name,
                "timestamp": str(int(datetime.now().timestamp())),
                "results": all_results
            }
            save_analysis(analysis_data)
        
        st.success("✅ Analysis Complete! Review your results below.")
        
        # Add save button (in case auto-save was skipped or failed)
        if st.button("💾 Save Analysis"):
            analysis_id = generate_analysis_id(molecule_name, disease_name)
            analysis_data = {
                "id": analysis_id,
                "molecule_name": molecule_name,
                "disease_name": disease_name,
                "timestamp": str(int(datetime.now().timestamp())),
                "results": all_results
            }
            save_analysis(analysis_data)
            st.success(f"✅ Analysis saved successfully!")
            st.rerun()
        
        # Create visualizations if in single molecule mode
        if analysis_mode == "Single Molecule":
            try:
                # Create the visualizations
                risk_fig = create_risk_radar(all_results)
                timeline_fig = create_timeline_gantt(all_results)
                market_fig = create_market_funnel(all_results)
                
                # Store figures in session state for later use in the dashboard tab
                st.session_state.visualizations = {
                    'risk_radar': risk_fig,
                    'timeline_gantt': timeline_fig,
                    'market_funnel': market_fig
                }
                
            except Exception as e:
                st.error(f"Error creating visualizations: {str(e)}")
                st.exception(e)
                st.session_state.visualizations = None
        
        # ======================================================================================
        # RESULTS SECTION WITH TABS
        # ======================================================================================
        tabs = st.tabs([
            "📊 Executive Summary", 
            "📈 Dashboard",
            "🔍 Detailed Analysis", 
            "⚠️ Risk Assessment", 
            "💬 Deep Dive",
            "📄 Download Report"
        ])
        
        # ==================== TAB 0: EXECUTIVE SUMMARY ====================
        with tabs[0]:
            if analysis_mode == "Single Molecule":
                st.markdown('<div class="glass-card">', unsafe_allow_html=True)
                st.markdown(f"### 🎯 Recommendation for {molecule_name}")
                
                # Display overall recommendation with confidence and uncertainty
                synthesis = all_results.get('synthesis', {})
                overall_rec = synthesis.get('recommendation', 'No recommendation available')
                confidence_display = synthesis.get('confidence_display', 'N/A')
                confidence_tooltip = synthesis.get('confidence_tooltip', '')
                needs_review = synthesis.get('needs_review', False)
                
                # Determine color and icon based on recommendation
                rec_upper = overall_rec.upper()
                if "PROCEED" in rec_upper and "CAUTION" not in rec_upper:
                    rec_color = "#4CAF50"  # Green
                    rec_icon = "✅"
                elif "CAUTION" in rec_upper or "REVIEW" in rec_upper:
                    rec_color = "#FFC107"  # Yellow
                    rec_icon = "⚠️"
                elif "NOT RECOMMENDED" in rec_upper or "REJECT" in rec_upper:
                    rec_color = "#F44336"  # Red
                    rec_icon = "❌"
                else:
                    rec_color = "#9E9E9E"  # Grey
                    rec_icon = "❓"
                
                # Recommendation with confidence interval
                rec_html = f"""
                    <div style="
                        background: rgba(30, 30, 30, 0.2);
                        border-radius: 10px;
                        padding: 1.2rem;
                        margin: 1rem 0;
                        border-left: 4px solid {rec_color};
                    ">
                        <div style="display: flex; align-items: center; gap: 0.8rem;">
                            <div style="font-size: 2rem; line-height: 1;">{rec_icon}</div>
                            <div>
                                <div style="font-size: 1.2rem; font-weight: 600; line-height: 1.4;">
                                    {overall_rec}
                                    <span style="font-size: 0.9rem; font-weight: normal; opacity: 0.8;" 
                                        title="{confidence_tooltip}">
                                        (Confidence: {confidence_display}) ⓘ
                                    </span>
                                </div>
                                {f'<div style="color: #FFA500; font-weight: bold; margin-top: 0.5rem;">'
                                f'    ⚠️ This recommendation requires additional review due to low confidence'
                                '</div>' if needs_review else ''}
                            </div>
                        </div>
                    </div>
                    """
                st.markdown(rec_html, unsafe_allow_html=True)
                
                # Agent confidence visualization
                agent_confidences = synthesis.get('agent_confidences', [])
                if agent_confidences:
                    st.markdown("### 🔍 Analysis Confidence")
                    st.markdown("Confidence levels from each analysis agent:")
                    
                    for agent in agent_confidences:
                        conf = agent['confidence']
                        weight = agent['weight']
                        agent_name = agent['agent']
                        
                        # Determine color based on confidence level
                        if conf >= 70:
                            color = "#4CAF50"  # Green
                        elif conf >= 40:
                            color = "#FFC107"  # Yellow
                        else:
                            color = "#F44336"  # Red
                        
                        # Create the confidence bar
                        st.markdown(f"""
                        <div style="margin-bottom: 0.5rem;">
                            <div style="display: flex; justify-content: space-between; margin-bottom: 0.2rem;">
                                <span>{agent_name}</span>
                                <span>{conf}%</span>
                            </div>
                            <div style="background: rgba(255, 255, 255, 0.1); height: 8px; border-radius: 4px; overflow: hidden;">
                                <div style="background: {color}; width: {conf}%; height: 100%;"></div>
                            </div>
                            <div style="font-size: 0.8rem; color: #aaa; text-align: right;">
                                Weight: {int(weight * 100)}%
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                
                # Add a small space before key factors
                st.markdown("<div style='margin-top: 1.5rem;'></div>", unsafe_allow_html=True)
                
                # Add Export to Excel button
                if st.button("💾 Export to Excel", key="export_excel_btn"):
                    with st.spinner("Preparing Excel export..."):
                        try:
                            excel_path = create_excel_export(
                                all_results, 
                                molecule_name, 
                                disease_name
                            )
                            
                            # Read the file and create a download button
                            with open(excel_path, "rb") as f:
                                excel_data = f.read()
                            
                            # Create a download button
                            st.download_button(
                                label="⬇️ Download Excel Report",
                                data=excel_data,
                                file_name=f"{molecule_name.replace(' ', '_')}_{disease_name.replace(' ', '_')}_analysis.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                help="Download a detailed Excel report with all analysis results"
                            )
                            
                            # Clean up the temporary file
                            try:
                                os.remove(excel_path)
                            except:
                                pass
                                
                        except Exception as e:
                            st.error(f"Error generating Excel export: {str(e)}")
                            st.exception(e)
                
                # Display key factors in a nice layout with icons
                key_factors = synthesis.get('key_factors', [])
                if key_factors:
                    st.markdown("### 📊 Key Factors")
                    cols = st.columns(2)
                    for i, factor in enumerate(key_factors):
                        with cols[i % 2]:
                            st.markdown(f"""
                            <div style="
                                background: rgba(255, 255, 255, 0.05);
                                border-radius: 8px;
                                padding: 12px;
                                margin-bottom: 10px;
                                border-left: 4px solid #4ECDC4;
                                font-size: 0.95rem;
                            ">
                                {factor}
                            </div>
                            """, unsafe_allow_html=True)
                
                
                # Add Find Similar Molecules section
                if 'similar_molecules' not in st.session_state:
                    st.session_state.similar_molecules = None
                
                if st.button("🔍 Find Similar Molecules", key="find_similar_btn"):
                    with st.spinner("Searching for similar molecules..."):
                        st.session_state.similar_molecules = find_similar_molecules(
                            molecule_name, disease_name
                        )
                
                if st.session_state.similar_molecules:
                    st.markdown("### 🧪 Similar Molecules")
                    st.markdown("Explore similar molecules that might be of interest:")
                    
                    # Create columns for the molecule cards
                    cols = st.columns(3)
                    for i, (mol, mol_type) in enumerate(st.session_state.similar_molecules):
                        with cols[i % 3]:
                            with st.container():
                                st.markdown(f"""
                                <div style="
                                    background: rgba(255, 255, 255, 0.1);
                                    border-radius: 10px;
                                    padding: 15px;
                                    margin-bottom: 15px;
                                    border-left: 4px solid #4ECDC4;
                                ">
                                    <h4>{mol}</h4>
                                    <p style="font-size: 0.9em; color: #ccc;">
                                        { {"same_class": "Same class/mechanism", "alternative": f"Alternative for {disease_name}"}.get(mol_type, "Related molecule") }
                                    </p>
                                    <button class="stButton" 
                                            onclick="analyzeMolecule('{mol}')" 
                                            style="width: 100%;">
                                        Analyze This
                                    </button>
                                </div>
                                """, unsafe_allow_html=True)
                    
                    # Add JavaScript to handle the Analyze This button
                    st.markdown("""
                    <script>
                    function analyzeMolecule(molecule) {
                        // Find the molecule input field and update its value
                        const inputs = document.querySelectorAll('input[type="text"]');
                        if (inputs.length > 0) {
                            inputs[0].value = molecule;
                            // Trigger change event
                            const event = new Event('input', {bubbles: true});
                            inputs[0].dispatchEvent(event);
                            
                            // Find and click the analyze button
                            const buttons = Array.from(document.querySelectorAll('button'));
                            const analyzeBtn = buttons.find(btn => 
                                btn.textContent.includes('Run Comprehensive Analysis')
                            );
                            if (analyzeBtn) analyzeBtn.click();
                        }
                    }
                    </script>
                    """, unsafe_allow_html=True)
                
                st.markdown('</div>', unsafe_allow_html=True)
        
            else:  # Compare Molecules
                st.markdown('<div class="glass-card">', unsafe_allow_html=True)
                st.markdown(f"### 🏆 Comparison Results for {disease_name}")
                
                # Prepare comparison data
                comparison_data = []
                for mol in molecules:
                    if mol in all_results:
                        result = all_results[mol]
                        synthesis = result.get('synthesis', {})
                        comparison_data.append({
                            'Molecule': mol,
                            'Recommendation': synthesis.get('recommendation', 'N/A'),
                            'Confidence': synthesis.get('confidence', 'N/A'),
                            'Key Strengths': '\n'.join(synthesis.get('strengths', [])),
                            'Key Weaknesses': '\n'.join(synthesis.get('weaknesses', [])),
                            'Overall Score': synthesis.get('score', 0)
                        })
                
                if comparison_data:
                    # Sort by score (highest first)
                    comparison_data.sort(key=lambda x: x['Overall Score'], reverse=True)
                    
                    # Create a styled dataframe
                    df = pd.DataFrame(comparison_data)
                    
                    # Function to apply color based on recommendation
                    def color_recommendation(val):
                        if 'recommend' in str(val).lower():
                            return 'background-color: #4CAF50; color: white;'
                        elif 'caution' in str(val).lower() or 'consider' in str(val).lower():
                            return 'background-color: #FFC107; color: black;'
                        elif 'not recommend' in str(val).lower():
                            return 'background-color: #F44336; color: white;'
                        return ''
                    
                    # Apply styling
                    styled_df = df.style.applymap(color_recommendation, subset=['Recommendation'])
                    
                    # Display the styled dataframe
                    st.dataframe(
                        styled_df,
                        column_config={
                            'Molecule': 'Molecule',
                            'Recommendation': 'Recommendation',
                            'Confidence': st.column_config.ProgressColumn(
                                'Confidence',
                                min_value=0,
                                max_value=100,
                                format='%d%%',
                            ),
                            'Key Strengths': 'Key Strengths',
                            'Key Weaknesses': 'Key Weaknesses',
                            'Overall Score': st.column_config.ProgressColumn(
                                'Overall Score',
                                min_value=0,
                                max_value=100,
                                format='%d',
                            ),
                        },
                        use_container_width=True,
                        hide_index=True,
                    )
                    
                    # Show winner
                    if len(comparison_data) > 1:
                        winner = comparison_data[0]
                        st.success(f"🏆 **Best Candidate:** {winner['Molecule']} (Score: {winner['Overall Score']}/100)")
                        if winner['Key Strengths']:
                            first_strength = winner['Key Strengths'].split('\n')[0]
                            st.markdown(f"**Why?** {first_strength}")
                        else:
                            st.markdown("**Why?** No strengths identified")
                
                st.markdown('</div>', unsafe_allow_html=True)

        # ==================== TAB 1: DASHBOARD ====================
        with tabs[1]:
            if analysis_mode == "Single Molecule" and hasattr(st.session_state, 'visualizations') and st.session_state.visualizations:
                st.markdown('<div class="glass-card">', unsafe_allow_html=True)
                st.markdown("### 📊 Interactive Dashboard")
                
                # Create two columns for the top row
                col1, col2 = st.columns(2)
                
                with col1:
                    st.plotly_chart(st.session_state.visualizations['risk_radar'], use_container_width=True)
                
                with col2:
                    st.plotly_chart(st.session_state.visualizations['market_funnel'], use_container_width=True)
                
                # Full width for the timeline
                st.plotly_chart(st.session_state.visualizations['timeline_gantt'], use_container_width=True)
                
                # Add export button
                if st.button("💾 Export Dashboard as HTML", use_container_width=True, key="export_dashboard"):
                    # Create a temporary file path
                    export_path = export_dashboard(
                        {
                            "risk_radar": st.session_state.visualizations['risk_radar'],
                            "market_funnel": st.session_state.visualizations['market_funnel'],
                            "timeline_gantt": st.session_state.visualizations['timeline_ganant']
                        },
                        filename=f"{molecule_name.replace(' ', '_')}_dashboard.html"
                    )
                    
                    # Create download link
                    with open(export_path, "rb") as f:
                        b64 = base64.b64encode(f.read()).decode()
                        href = f'<a href="data:file/html;base64,{b64}" download="{os.path.basename(export_path)}">Download Dashboard</a>'
                        st.markdown(href, unsafe_allow_html=True)
                
                st.markdown('</div>', unsafe_allow_html=True)
            elif analysis_mode != "Single Molecule":
                st.info("Dashboard view is only available in Single Molecule mode.")
            else:
                st.warning("No visualization data available. Please run the analysis first.")

        # ==================== TAB 2: DETAILED ANALYSIS ====================
        with tabs[2]:
            st.markdown('<div class="glass-card">', unsafe_allow_html=True)
            st.markdown("### 🔍 Detailed Analysis")
            
            if analysis_mode == "Single Molecule":
                # Agent sections
                agent_sections = {
                    "patent_analysis": "📜 Patent Analysis",
                    "clinical_analysis": "🏥 Clinical Trials",
                    "market_analysis": "📈 Market Analysis",
                    "safety_analysis": "⚠️ Safety Profile"
                }
                
                for agent_key, agent_title in agent_sections.items():
                    if agent_key in all_results:
                        agent_data = all_results[agent_key]
                        
                        with st.expander(agent_title, expanded=False):
                            if isinstance(agent_data, dict) and 'findings' in agent_data:
                                for finding in agent_data['findings']:
                                    # Display finding with evidence badge
                                    col1, col2 = st.columns([4, 1])
                                    with col1:
                                        st.markdown(f"**Finding:** {finding.get('finding', '')}")
                                        if 'implications' in finding:
                                            st.markdown(f"*Implications:* {finding['implications']}")
                                        if 'recommendation' in finding:
                                            st.markdown(f"*Recommendation:* {finding['recommendation']}")
                                        
                                        # Display sources if available
                                        if 'sources' in finding and finding['sources']:
                                            with st.expander("View Sources", expanded=False):
                                                for src in finding['sources']:
                                                    st.markdown(f"- {src.get('type', 'Source').title()}: {src.get('url', 'No URL')}")
                                    
                                    with col2:
                                        # Display evidence strength badge
                                        score = finding.get('evidence_strength', 0)
                                        if hasattr(master_agent, 'get_evidence_badge'):
                                            st.markdown(
                                                master_agent.get_evidence_badge(score),
                                                unsafe_allow_html=True
                                            )
                                    
                                    st.markdown("---")
                            else:
                                # Fallback for agents without structured findings
                                conf = agent_data.get("confidence", "N/A")
                                st.write(f"**Confidence:** {conf}")
                                if "error" in agent_data:
                                    st.error(f"Error: {agent_data['error']}")
                                else:
                                    for k, v in agent_data.items():
                                        if k != "confidence":
                                            st.write(f"**{k}:** {v}")
            
            else:  # Compare Molecules
                for mol in molecules:
                    if mol in all_results:
                        result = all_results[mol]
                        with st.expander(f"📊 {mol}", expanded=False):
                            for agent_name, agent_result in result.items():
                                if agent_name in master_agent.agents:
                                    with st.expander(agent_name.replace("_", " ").title(), expanded=False):
                                        conf = agent_result.get("confidence", "N/A")
                                        st.write(f"**Confidence:** {conf}")
                                        if "error" in agent_result:
                                            st.error(f"Error: {agent_result['error']}")
                                        else:
                                            for k, v in agent_result.items():
                                                if k != "confidence":
                                                    st.write(f"**{k}:** {v}")
            
            st.markdown('</div>', unsafe_allow_html=True)

        # ==================== TAB 3: RISK ASSESSMENT ====================
        with tabs[3]:
            st.markdown('<div class="glass-card">', unsafe_allow_html=True)
            
            if analysis_mode == "Single Molecule":
                synthesis = all_results.get('synthesis', {})
                risks = synthesis.get("risks", [])
                
                if risks and risks != ["No major risks identified"]:
                    st.markdown("### ⚠️ Identified Risks")
                    for risk in risks:
                        st.markdown(f"🔴 {risk}")
                else:
                    st.success("✅ No significant risks identified")
                    
            else:  # Compare Molecules
                for mol in molecules:
                    if mol in all_results:
                        result = all_results[mol]
                        with st.expander(f"⚠️ Risks for {mol}", expanded=False):
                            risks = []
                            for agent_name, agent_result in result.items():
                                if isinstance(agent_result, dict) and "risks" in agent_result:
                                    risks.extend(agent_result["risks"])
                            
                            if risks:
                                st.markdown("### ⚠️ Identified Risks")
                                for r in risks:
                                    st.markdown(f"- ⚠️ {r}")
                            else:
                                st.success("✅ No significant risks identified")
            
            st.markdown("</div>", unsafe_allow_html=True)

        # ==================== TAB 4: DEEP DIVE ====================
        with tabs[4]:
            st.markdown('<div class="glass-card">', unsafe_allow_html=True)
            st.markdown("### 🔍 Deep Dive Analysis")
            
            # Initialize session state for chat if not exists
            if 'chat_history' not in st.session_state:
                st.session_state.chat_history = []
            
            # Agent selection
            if hasattr(master_agent, 'agent_display_names'):
                agent_options = list(master_agent.agent_display_names.values())
            else:
                agent_options = [name.replace("_", " ").title() for name in master_agent.agents.keys()]
            
            selected_agent = st.selectbox(
                "Select an agent to chat with:",
                options=agent_options,
                key="selected_agent"
            )
            
            # Get the internal agent name from display name
            if hasattr(master_agent, 'agent_display_names'):
                agent_name = next(
                    (k for k, v in master_agent.agent_display_names.items() 
                     if v == selected_agent),
                    None
                )
            else:
                agent_name = selected_agent.lower().replace(" ", "_")
            
            # Display chat history
            chat_container = st.container()
            with chat_container:
                for msg in st.session_state.chat_history:
                    with st.chat_message(msg["role"], avatar=msg.get("avatar", None)):
                        if msg.get('agent'):
                            st.markdown(f"**{msg.get('agent', '')}**")
                        st.markdown(msg["content"])
            
            # Chat input
            if prompt := st.chat_input("Ask a question about the analysis..."):
                # Add user message to chat
                st.session_state.chat_history.append({
                    "role": "user",
                    "content": prompt,
                    "agent": None
                })
                
                # Get agent response
                with st.spinner(f"{selected_agent} is thinking..."):
                    if hasattr(master_agent, 'query_agent'):
                        response = asyncio.run(
                            master_agent.query_agent(
                                molecule=molecule_name if analysis_mode == "Single Molecule" else None,
                                disease=disease_name,
                                agent_name=agent_name,
                                question=prompt
                            )
                        )
                    else:
                        response = "This feature requires the query_agent method in MasterAgent."
                
                # Add assistant response to chat
                st.session_state.chat_history.append({
                    "role": "assistant",
                    "content": response,
                    "agent": selected_agent,
                    "avatar": "🤖"
                })
                
                # Rerun to update the chat display
                st.rerun()
            
            st.markdown('</div>', unsafe_allow_html=True)

        # ==================== TAB 5: REPORT DOWNLOAD ====================
        with tabs[5]:
            st.markdown('<div class="glass-card">', unsafe_allow_html=True)
            pdf_file = all_results.get("pdf_report")
            if pdf_file:
                st.markdown("### 📄 Download Comprehensive Report")
                st.markdown("Your detailed analysis report is ready for download.")
                try:
                    with open(pdf_file, "rb") as f:
                        st.download_button(
                            label="📥 Download PDF Report",
                            data=f,
                            file_name=f"{molecule_name}_{disease_name}_analysis.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
                except Exception as e:
                    st.error(f"Error loading PDF: {e}")
            else:
                st.warning("PDF report not available.")
            st.markdown("</div>", unsafe_allow_html=True)

# ======================================================================================
# FOOTER
# ======================================================================================
st.markdown("---")
st.markdown(
    """
    <div style="text-align: center; color: rgba(255, 255, 255, 0.5); padding: 20px;">
        <p>IntelliDrug AI © 2025 | Powered by Multi-Agent Systems & Google Gemini</p>
        <p style="font-size: 0.9rem;">Accelerating pharmaceutical innovation through artificial intelligence</p>
    </div>
    """,
    unsafe_allow_html=True
)